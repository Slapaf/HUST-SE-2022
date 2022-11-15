from pathlib import Path

from flask import Flask, Response
import os
import zipfile
import mimetypes
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from init import app


@app.route('/download/<path:filename>')
def download_file(filename):
    """下载文件

    Args:
        filename (str): 文件名

    """
    base_dir = os.getcwd()  # ! 目前代表该 py 文件所在目录，需要修改为收集所在目录
    path_name = os.path.join(base_dir, filename)
    f = open(path_name, "rb")
    response = Response(f.readlines())
    mime_type = mimetypes.guess_type(filename)[0]
    response.headers['Content-Type'] = mime_type
    response.headers['Content-Disposition'] = 'attachment; filename={}'.format(
        filename.encode().decode('latin-1')
    )
    return response


def generate_excel(due_list: list,
                   submitted_list: list,
                   excel_name: str = 'demo',
                   excel_path: str = './'):
    """生成汇总 Excel 表格

    Args:
        due_list (list): 应交名单列表
        submitted_list (list): 已交名单列表
        excel_name (str, optional): 汇总表格名称，默认 'demo'
        excel_path (str, optional): 汇总表格地址. 默认 './'.
    """
    submit_status = [
        '已提交' if name in submitted_list else '未提交' for name in due_list
    ]
    # excel_path += excel_name
    excel_path = os.path.join(excel_path, excel_name)
    df = pd.DataFrame({'成员名单': due_list, '提交状态': submit_status})
    df.to_excel(excel_path, sheet_name='提交情况', index=False)
    # 设置样式
    # red_fill = PatternFill('solid', start_color='FF9198')
    # green_fill = PatternFill("solid", start_color='A1DDAA')
    # # 应用样式
    # wb = load_workbook(excel_path)
    # ws = wb.active
    # cells = ws.iter_rows(min_row=2, min_col=2)
    # for cell in cells:
    #     cell[0].fill = green_fill if cell[0].value == '已提交' else red_fill
    # # 保存表格
    # wb.save(excel_name)
    # wb.close()


def make_response(file_path: str, file_name: str) -> Response:
    """生成返回文件的 HTTP 头部

    Args:
        file_path: 文件所在路径
        file_name: 文件名

    Returns:
        请求
    """
    f = open(os.path.join(file_path, file_name), 'rb')
    response = Response(f.readlines())
    mime_type = mimetypes.guess_type(file_name)[0]
    response.headers['Content-Type'] = mime_type
    response.headers['Content-Disposition'] = 'attachment; filename={}'.format(
        file_name.encode().decode('latin-1')
    )
    return response


def generate_zip(to_zip: str, save_zip_name: str):
    """压缩文件及文件夹

    save_zip_name 带目录，若不带目录就是当前目录

    Args:
        to_zip (str): 待压缩的目录
        save_zip_name (str): 压缩文件名
    """
    # * 1.先判断输出 save_zip_name 的上级是否存在(判断绝对目录)，否则创建目录
    save_zip_dir = os.path.split(
        os.path.abspath(save_zip_name))[0]  # save_zip_name的上级目录
    print(save_zip_dir)
    if not os.path.exists(save_zip_dir):
        os.makedirs(save_zip_dir)
        print('创建新目录 %s' % save_zip_dir)
    f = zipfile.ZipFile(os.path.abspath(save_zip_name), 'w',
                        zipfile.ZIP_DEFLATED)
    # * 2.判断要被压缩的 to_zip 是否目录还是文件，是目录就遍历操作，是文件直接压缩
    if not os.path.isdir(os.path.abspath(to_zip)):  # 如果不是目录,那就是文件
        if os.path.exists(os.path.abspath(to_zip)):  # 判断文件是否存在
            f.write(to_zip)
            f.close()
            print('%s 压缩为 %s' % (to_zip, save_zip_name))
        else:
            print('%s 文件不存在' % os.path.abspath(to_zip))
    else:
        if os.path.exists(os.path.abspath(to_zip)):  # 判断目录是否存在，遍历目录
            zip_list = []
            for cur_dir, sub_dirs, files in os.walk(to_zip):  # 遍历目录，加入列表
                for fileItem in files:
                    zip_list.append(os.path.join(cur_dir, fileItem))
                for dirItem in sub_dirs:
                    zip_list.append(os.path.join(cur_dir, dirItem))
            # 读取列表压缩目录和文件
            for i in zip_list:
                # replace是减少压缩文件的一层目录，即压缩文件不包括to_zip这个目录
                f.write(i, i.replace(to_zip, ''))
                # print('%s压缩到%s'%(i,save_zip_name))
            f.close()
            print('%s 压缩为 %s' % (to_zip, save_zip_name))
        else:
            print('%s 文件夹不存在' % os.path.abspath(to_zip))
